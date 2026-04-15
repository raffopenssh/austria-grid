#!/usr/bin/env python3
"""Parse AustroControl obstacle spreadsheet into JSON."""

import json
import sys
import openpyxl
from collections import Counter

INPUT_FILE = '/tmp/shelley-screenshots/upload_1e11474ec7de0285.xlsx'
OUTPUT_FILE = '/home/exedev/austria-grid/data/obstacles_all.json'
SHEET_NAME = 'Alle - All'


def parse_bool(val):
    """Parse 'ja / yes' -> True, 'nein / no' -> False, '---' -> None."""
    if val is None or val.strip() == '---':
        return None
    v = val.strip().lower()
    if v.startswith('ja') or v == 'yes':
        return True
    if v.startswith('nein') or v == 'no':
        return False
    return None


def parse_height_m(val):
    """Parse '427 / 1400' -> 427.0 (meters), '---' -> None.
    Also handles '1378 * / 4521 *' (footnote marker).
    """
    if val is None:
        return None
    val = val.strip()
    if val == '---':
        return None
    # Take the part before ' / '
    parts = val.split('/')
    m_str = parts[0].strip().replace('*', '').strip()
    try:
        return float(m_str)
    except ValueError:
        return None


def parse_type(art_str):
    """Parse 'Windpark / Windmill farm' -> ('Windpark', 'Windmill farm')."""
    if not art_str:
        return (None, None)
    parts = art_str.split(' / ', 1)
    if len(parts) == 2:
        return (parts[0].strip(), parts[1].strip())
    return (art_str.strip(), None)


def parse_geometry_type(geom_str):
    """Parse geometry string to a normalized type."""
    if not geom_str:
        return 'unknown'
    g = geom_str.lower()
    if 'point (grouped)' in g or 'punkt (gruppiert)' in g:
        return 'point_grouped'
    if 'curve (grouped)' in g or 'linie (gruppiert)' in g:
        return 'curve_grouped'
    if 'surface' in g or 'fläche' in g:
        return 'surface'
    if 'curve' in g or 'linie' in g:
        return 'curve'
    if 'point' in g or 'punkt' in g:
        return 'point'
    return 'unknown'


def split_multiline(val):
    """Split a multi-line cell value, filtering out empty lines."""
    if val is None:
        return []
    return [line.strip() for line in str(val).split('\n') if line.strip()]


def split_multiline_keep_blanks(val):
    """Split a multi-line cell value, keeping structure (for coords with blank separator lines)."""
    if val is None:
        return []
    return [line.strip() for line in str(val).split('\n')]


def parse_coord_pairs(coord_str):
    """Parse decimal degree coordinate pairs from multi-line string.
    Format: 'lat lon\nlat lon\n...'
    May have blank lines between pairs (for curves).
    Returns list of (lat, lon) tuples.
    """
    if not coord_str:
        return []
    pairs = []
    for line in str(coord_str).split('\n'):
        line = line.strip()
        if not line:
            continue
        parts = line.split()
        if len(parts) >= 2:
            try:
                lat = float(parts[0])
                lon = float(parts[1])
                pairs.append((lat, lon))
            except ValueError:
                pass
    return pairs


def parse_horizontal_extent(val):
    """Parse '50 x 20' -> string, '---' -> None."""
    if val is None:
        return None
    val = str(val).strip()
    if val == '---':
        return None
    return val


def parse_radius(val):
    """Parse radius in meters. '---' -> None, numeric -> float."""
    if val is None:
        return None
    val = str(val).strip()
    if val == '---':
        return None
    try:
        return float(val)
    except ValueError:
        return None


def main():
    wb = openpyxl.load_workbook(INPUT_FILE, data_only=True)
    ws = wb[SHEET_NAME]

    obstacles = []
    type_counter = Counter()
    geom_counter = Counter()
    errors = []

    for row_idx in range(4, ws.max_row + 1):
        # Read all columns (1-indexed in openpyxl)
        def cell(col_0):
            return ws.cell(row_idx, col_0 + 1).value

        region = cell(0)
        # Skip empty rows
        if not region:
            continue

        district = cell(1)
        location = cell(2)
        art = cell(3)
        geom_raw = cell(4)
        # col 5 = DMS coords (skip, use decimal)
        coord_decimal = cell(6)
        # col 7 = vertical ref system (skip)
        elev_raw = cell(8)
        height_agl_raw = cell(9)
        day_marking_raw = cell(10)
        lighted_raw = cell(11)
        # col 12 = data quality (skip)
        horiz_extent_raw = cell(13)
        horiz_radius_raw = cell(14)
        # cols 15-17 = accuracy (skip)
        identifier = cell(18)

        type_de, type_en = parse_type(art)
        geometry_type = parse_geometry_type(geom_raw)
        type_counter[type_en or type_de or 'Unknown'] += 1
        geom_counter[geometry_type] += 1

        # Parse coordinate pairs
        coord_pairs = parse_coord_pairs(coord_decimal)
        n_coords = len(coord_pairs)

        if n_coords == 0:
            errors.append(f"Row {row_idx}: No coordinates found for {location}")
            continue

        # Parse multi-line height/marking fields
        elev_lines = split_multiline(elev_raw)
        height_lines = split_multiline(height_agl_raw)
        day_lines = split_multiline(day_marking_raw)
        lighted_lines = split_multiline(lighted_raw)

        # Build points array
        # For curves/curve_grouped: heights have entries for endpoints AND spans
        # (n_coords endpoints + n_coords-1 spans = 2*n_coords - 1 lines)
        # For point_grouped/point: heights match 1:1 with coords
        # For surface: single height value for all boundary points

        points = []

        if geometry_type in ('curve', 'curve_grouped'):
            # Curves: height lines alternate: endpoint, span, endpoint, span, ...
            # So endpoint heights are at indices 0, 2, 4, ... (every other)
            # n_coords endpoints -> indices 0, 2, ..., 2*(n_coords-1)
            expected_height_lines = 2 * n_coords - 1 if n_coords > 1 else 1

            for i, (lat, lon) in enumerate(coord_pairs):
                height_idx = i * 2  # endpoint indices: 0, 2, 4, ...

                elev_m = parse_height_m(elev_lines[height_idx]) if height_idx < len(elev_lines) else None
                hgt_m = parse_height_m(height_lines[height_idx]) if height_idx < len(height_lines) else None
                dm = parse_bool(day_lines[height_idx]) if height_idx < len(day_lines) else None
                lit = parse_bool(lighted_lines[height_idx]) if height_idx < len(lighted_lines) else None

                points.append({
                    'lat': round(lat, 8),
                    'lon': round(lon, 8),
                    'elev_m': elev_m,
                    'height_agl_m': hgt_m,
                    'day_marked': dm,
                    'lighted': lit,
                })

            # Also capture span data as separate entries between endpoints
            # (the cable heights between towers are important for aviation)
            for i in range(n_coords - 1):
                span_idx = i * 2 + 1
                if span_idx < len(elev_lines):
                    elev_m = parse_height_m(elev_lines[span_idx])
                    hgt_m = parse_height_m(height_lines[span_idx]) if span_idx < len(height_lines) else None
                    dm = parse_bool(day_lines[span_idx]) if span_idx < len(day_lines) else None
                    lit = parse_bool(lighted_lines[span_idx]) if span_idx < len(lighted_lines) else None

                    # Midpoint of the two endpoints
                    lat1, lon1 = coord_pairs[i]
                    lat2, lon2 = coord_pairs[i + 1]
                    mid_lat = round((lat1 + lat2) / 2, 8)
                    mid_lon = round((lon1 + lon2) / 2, 8)

                    points.append({
                        'lat': mid_lat,
                        'lon': mid_lon,
                        'elev_m': elev_m,
                        'height_agl_m': hgt_m,
                        'day_marked': dm,
                        'lighted': lit,
                        'is_span': True,
                    })

        elif geometry_type == 'surface':
            # Surface: single height for all boundary points
            elev_m = parse_height_m(elev_lines[0]) if elev_lines else None
            hgt_m = parse_height_m(height_lines[0]) if height_lines else None
            dm = parse_bool(day_lines[0]) if day_lines else None
            lit = parse_bool(lighted_lines[0]) if lighted_lines else None

            for lat, lon in coord_pairs:
                points.append({
                    'lat': round(lat, 8),
                    'lon': round(lon, 8),
                    'elev_m': elev_m,
                    'height_agl_m': hgt_m,
                    'day_marked': dm,
                    'lighted': lit,
                })

        else:
            # point or point_grouped: 1:1 mapping
            for i, (lat, lon) in enumerate(coord_pairs):
                elev_m = parse_height_m(elev_lines[i]) if i < len(elev_lines) else None
                hgt_m = parse_height_m(height_lines[i]) if i < len(height_lines) else None
                dm = parse_bool(day_lines[i]) if i < len(day_lines) else None
                lit = parse_bool(lighted_lines[i]) if i < len(lighted_lines) else None

                points.append({
                    'lat': round(lat, 8),
                    'lon': round(lon, 8),
                    'elev_m': elev_m,
                    'height_agl_m': hgt_m,
                    'day_marked': dm,
                    'lighted': lit,
                })

        # Compute centroid
        all_lats = [p['lat'] for p in points if not p.get('is_span')]
        all_lons = [p['lon'] for p in points if not p.get('is_span')]
        if not all_lats:  # fallback
            all_lats = [p['lat'] for p in points]
            all_lons = [p['lon'] for p in points]
        centroid_lat = round(sum(all_lats) / len(all_lats), 8)
        centroid_lon = round(sum(all_lons) / len(all_lons), 8)

        # Max elevation and height AGL
        elevs = [p['elev_m'] for p in points if p['elev_m'] is not None]
        hgts = [p['height_agl_m'] for p in points if p['height_agl_m'] is not None]
        max_elev = max(elevs) if elevs else None
        max_hgt = max(hgts) if hgts else None

        obstacle = {
            'id': str(identifier).strip() if identifier else None,
            'region': str(region).strip() if region else None,
            'district': str(district).strip() if district else None,
            'location': str(location).strip() if location else None,
            'type_de': type_de,
            'type_en': type_en,
            'geometry_type': geometry_type,
            'points': points,
            'centroid_lat': centroid_lat,
            'centroid_lon': centroid_lon,
            'max_elev_m': max_elev,
            'max_height_agl_m': max_hgt,
            'horizontal_extent': parse_horizontal_extent(horiz_extent_raw),
            'horizontal_radius_m': parse_radius(horiz_radius_raw),
        }
        obstacles.append(obstacle)

    # Write output
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        json.dump(obstacles, f, ensure_ascii=False, indent=2)

    # Report
    print(f"Total obstacles parsed: {len(obstacles)}")
    print(f"Output: {OUTPUT_FILE}")
    print(f"\nGeometry type distribution:")
    for gt, count in geom_counter.most_common():
        print(f"  {gt}: {count}")
    print(f"\nObstacle type distribution:")
    for t, count in type_counter.most_common():
        print(f"  {t}: {count}")
    if errors:
        print(f"\nErrors ({len(errors)}):")
        for e in errors[:20]:
            print(f"  {e}")

    # Some basic validation
    print(f"\n--- Validation ---")
    no_elev = sum(1 for o in obstacles if o['max_elev_m'] is None)
    no_hgt = sum(1 for o in obstacles if o['max_height_agl_m'] is None)
    print(f"  Obstacles with no elevation: {no_elev}")
    print(f"  Obstacles with no AGL height: {no_hgt}")

    # Check a sample
    for o in obstacles:
        if 'Windpark Zagersdorf' in (o['location'] or ''):
            print(f"\n--- Sample: {o['location']} ---")
            print(f"  Type: {o['type_en']}")
            print(f"  Geometry: {o['geometry_type']}")
            print(f"  Points: {len(o['points'])}")
            for p in o['points']:
                print(f"    {p}")
            print(f"  Centroid: ({o['centroid_lat']}, {o['centroid_lon']})")
            print(f"  Max elev: {o['max_elev_m']}m, Max AGL: {o['max_height_agl_m']}m")
            break

    for o in obstacles:
        if 'Fiderepasshütte' in (o['location'] or ''):
            print(f"\n--- Sample: {o['location']} ---")
            print(f"  Type: {o['type_en']}")
            print(f"  Geometry: {o['geometry_type']}")
            print(f"  Points: {len(o['points'])}")
            for p in o['points']:
                print(f"    {p}")
            print(f"  Max elev: {o['max_elev_m']}m, Max AGL: {o['max_height_agl_m']}m")
            break


if __name__ == '__main__':
    main()
